from __future__ import annotations

from datetime import date, datetime, time, timedelta
from typing import Any, cast

import pytz
from django.db import models
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.core.selectors import build_visible_user_ids
from core.models import ClockSession, ProjectBudget, Transaction
from core.permissions import IsExecutive, IsManager
from tables import Project, ProjectAssignment, Tag as TagModel, Task, TimeEntry, User


def _get_query_str(request: Request, key: str, default: str | None = None) -> str | None:
    value = request.query_params.get(key, default)
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
    else:
        value = str(value)
    return value or None


class ReportsSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            period = request.query_params.get("period", "week")
            user_id = request.query_params.get("user_id")
            project = request.query_params.get("project")
            start_date = request.query_params.get("start_date")
            end_date = request.query_params.get("end_date")
            export = request.query_params.get("export")
            user = request.user
            qs = TimeEntry.objects.all()

            # Employees are hard-restricted to their own entries. They cannot
            # override this via the user_id query param.
            is_employee_only = (
                not user.is_superuser
                and getattr(user, "role", "") not in ("superuser", "manager")
            )
            if is_employee_only:
                qs = qs.filter(user_id=user.id)
            elif not (user.is_superuser or user.role == "superuser"):
                visible_user_ids = build_visible_user_ids(user)
                visible_entry_ids = TimeEntry.objects.filter(visible_to=user).values("id")
                qs = qs.filter(Q(user_id__in=visible_user_ids) | Q(id__in=visible_entry_ids))
            if user_id and not is_employee_only:
                if user_id == "me":
                    qs = qs.filter(user_id=user.id)
                else:
                    qs = qs.filter(user_id=user_id)
            if project:
                qs = qs.filter(project_id=project)

            if start_date:
                tz = pytz.timezone("Asia/Kolkata")
                try:
                    s_date = datetime.strptime(start_date, "%Y-%m-%d").date()
                    s_dt = tz.localize(datetime.combine(s_date, time.min))
                    qs = qs.filter(start_datetime__gte=s_dt)
                except ValueError:
                    pass
            if end_date:
                tz = pytz.timezone("Asia/Kolkata")
                try:
                    e_date = datetime.strptime(end_date, "%Y-%m-%d").date()
                    e_dt = tz.localize(datetime.combine(e_date, time.max))
                    qs = qs.filter(start_datetime__lte=e_dt)
                except ValueError:
                    pass

            if export in ("xlsx", "pdf"):
                max_export = 10000
                if qs.count() > max_export:
                    return Response({"detail": f"Export limited to {max_export} entries. Please filter your query."}, status=400)

                entries = qs.select_related("user", "project", "task").values(
                    "user__first_name",
                    "user__username",
                    "project__name",
                    "task__title",
                    "start_datetime",
                    "end_datetime",
                    "duration",
                ).order_by("start_datetime")[:max_export]

                tz = pytz.timezone("Asia/Kolkata")
                processed_entries = []
                for row in entries:
                    start_local = row["start_datetime"].astimezone(tz) if row["start_datetime"] else None
                    end_local = row["end_datetime"].astimezone(tz) if row["end_datetime"] else None
                    if start_local:
                        start_local = start_local.replace(tzinfo=None)
                    if end_local:
                        end_local = end_local.replace(tzinfo=None)
                    processed_entries.append({**row, "start_local": start_local, "end_local": end_local})

                if export == "xlsx":
                    try:
                        from openpyxl import Workbook
                    except ImportError:
                        return Response({"detail": "XLSX generation not available on server."}, status=501)

                    wb = Workbook()
                    ws = wb.active
                    if ws is None:
                        ws = wb.create_sheet("Time Entries")
                    else:
                        ws.title = "Time Entries"

                    ws.append(["User", "Project", "Task", "Start Time", "Stop Time", "Total Hours"])
                    for row in processed_entries:
                        user_name = row["user__first_name"] or row["user__username"] or "Unknown"
                        project_name = row["project__name"] or "No Project"
                        task_title = row["task__title"] or ""
                        total_hours = round((row["duration"].total_seconds() / 3600.0) if row["duration"] else 0, 2)
                        ws.append([user_name, project_name, task_title, row["start_local"], row["end_local"], total_hours])

                    for ws_row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
                        for cell in ws_row:
                            cell.number_format = "YYYY-MM-DD HH:MM:SS"

                    from django.http import HttpResponse

                    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    response["Content-Disposition"] = 'attachment; filename="time_entries.xlsx"'
                    wb.save(response)
                    return response

                from io import BytesIO

                from django.http import HttpResponse

                try:
                    from reportlab.lib.pagesizes import letter
                    from reportlab.lib.units import inch
                    from reportlab.pdfgen import canvas
                except Exception:
                    return Response({"detail": "PDF generation not available on server."}, status=501)

                response = HttpResponse(content_type="application/pdf")
                response["Content-Disposition"] = 'attachment; filename="time_entries.pdf"'
                buffer = BytesIO()
                pdf_canvas = canvas.Canvas(buffer, pagesize=letter)
                _, height = letter
                y = height - 1 * inch
                headers = ["User", "Project", "Task", "Start Time", "Stop Time", "Total Hours"]
                col_widths = [1.2 * inch, 1.5 * inch, 1.5 * inch, 1.2 * inch, 1.2 * inch, 1 * inch]

                pdf_canvas.setFont("Helvetica-Bold", 10)
                x = 0.5 * inch
                for i, header in enumerate(headers):
                    pdf_canvas.drawString(x, y, header)
                    x += col_widths[i]
                y -= 20
                pdf_canvas.setFont("Helvetica", 9)

                for row in processed_entries:
                    if y < 1 * inch:
                        pdf_canvas.showPage()
                        y = height - 1 * inch
                        pdf_canvas.setFont("Helvetica-Bold", 10)
                        x = 0.5 * inch
                        for i, header in enumerate(headers):
                            pdf_canvas.drawString(x, y, header)
                            x += col_widths[i]
                        y -= 20
                        pdf_canvas.setFont("Helvetica", 9)

                    user_name = row["user__first_name"] or row["user__username"] or "Unknown"
                    project_name = row["project__name"] or "No Project"
                    task_title = row["task__title"] or ""
                    start_time = str(row["start_local"])
                    stop_time = str(row["end_local"])
                    total_hours = f"{round((row['duration'].total_seconds() / 3600.0) if row['duration'] else 0, 2)} h"
                    row_data = [user_name, project_name, task_title, start_time, stop_time, total_hours]

                    x = 0.5 * inch
                    for i, cell in enumerate(row_data):
                        pdf_canvas.drawString(x, y, str(cell)[:20])
                        x += col_widths[i]
                    y -= 15

                pdf_canvas.save()
                response.write(buffer.getvalue())
                buffer.close()
                return response

            if period == "totals":
                tz = pytz.timezone("Asia/Kolkata")
                now_ist = timezone.now().astimezone(tz)
                today_ist = now_ist.date()
                week_start_ist = today_ist - timedelta(days=today_ist.weekday())
                month_start_ist = today_ist.replace(day=1)

                today_start_dt = tz.localize(datetime.combine(today_ist, time.min))
                today_end_dt = tz.localize(datetime.combine(today_ist, time.max))
                week_start_dt = tz.localize(datetime.combine(week_start_ist, time.min))
                month_start_dt = tz.localize(datetime.combine(month_start_ist, time.min))

                def calc_hours(queryset):
                    try:
                        total_obj = queryset.distinct().aggregate(total=models.Sum("duration"))
                        total = total_obj.get("total")
                        return round((total.total_seconds() / 3600.0) if total else 0.0, 2)
                    except Exception:
                        return 0.0

                return Response(
                    {
                        "today": calc_hours(qs.filter(start_datetime__range=(today_start_dt, today_end_dt))),
                        "week": calc_hours(qs.filter(start_datetime__gte=week_start_dt)),
                        "month": calc_hours(qs.filter(start_datetime__gte=month_start_dt)),
                    }
                )

            tz = pytz.timezone("Asia/Kolkata")
            data = qs.annotate(day=TruncDate("start_datetime", tzinfo=tz)).values("day").annotate(total=models.Sum("duration")).order_by("day")
            out = []
            for row in data:
                try:
                    total_seconds = row["total"].total_seconds() if row["total"] else 0
                    day_date = row["day"]
                    if isinstance(day_date, date):
                        out.append({"date": day_date.isoformat(), "hours": round(total_seconds / 3600.0, 2)})
                except Exception:
                    continue

            if request.query_params.get("format") == "csv":
                import csv
                from django.http import HttpResponse

                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = 'attachment; filename="summary.csv"'
                writer = csv.writer(response)
                writer.writerow(["date", "hours"])
                for row in out:
                    writer.writerow([row["date"], row["hours"]])
                return response

            return Response(out)
        except Exception:
            return Response({"detail": "An error occurred while generating the report."}, status=500)


class TeamSummaryReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        user = cast(User, request.user)
        qp_user_id = request.query_params.get("user_id")
        qp_project_id = request.query_params.get("project_id") or request.query_params.get("project")
        qp_billable = request.query_params.get("billable")
        start_date_param = _get_query_str(request, "start_date")
        end_date_param = _get_query_str(request, "end_date")
        # Day-wise filter: a single `day` collapses the window to that date and
        # takes precedence over any start/end range.
        day_param = _get_query_str(request, "day")
        if day_param:
            start_date_param = day_param
            end_date_param = day_param

        tz = pytz.timezone("Asia/Kolkata")
        now_ist = timezone.now().astimezone(tz)
        default_start_date = now_ist.replace(day=1).date()
        default_end_date = (default_start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        start_date = default_start_date
        end_date = default_end_date
        if start_date_param:
            try:
                start_date = datetime.fromisoformat(start_date_param).date()
            except ValueError:
                return Response({"detail": "Invalid start_date format"}, status=400)
        if end_date_param:
            try:
                end_date = datetime.fromisoformat(end_date_param).date()
            except ValueError:
                return Response({"detail": "Invalid end_date format"}, status=400)

        start_dt = tz.localize(datetime.combine(start_date, time.min))
        end_dt = tz.localize(datetime.combine(end_date, time.max))

        qs = TimeEntry.objects.all()
        # Employees are hard-restricted to their own time entries regardless of
        # any user_id query param. Managers/superusers still see their visible set.
        is_employee_only = (
            not user.is_superuser
            and getattr(user, "role", "") not in ("superuser", "manager")
        )
        if is_employee_only:
            qs = qs.filter(user_id=user.id)
        elif not (user.is_superuser or user.role == "superuser"):
            visible_user_ids = build_visible_user_ids(user)
            visible_entry_ids = TimeEntry.objects.filter(visible_to=user).values("id")
            qs = qs.filter(Q(user_id__in=visible_user_ids) | Q(id__in=visible_entry_ids))
        if qp_user_id and not is_employee_only:
            if qp_user_id == "me":
                qs = qs.filter(user_id=user.id)
            else:
                qs = qs.filter(user_id=qp_user_id)
        if qp_project_id:
            qs = qs.filter(project_id=qp_project_id)
        if qp_billable is not None:
            if qp_billable.lower() == "true":
                qs = qs.filter(billable=True)
            elif qp_billable.lower() == "false":
                qs = qs.filter(billable=False)

        qs = qs.filter(start_datetime__range=(start_dt, end_dt))

        if user.is_superuser or user.role == "superuser":
            users_qs = User.objects.filter(is_active=True).exclude(role="client")
        elif user.role == "manager":
            users_qs = User.objects.filter(id__in=build_visible_user_ids(user), is_active=True).exclude(role="client")
        else:
            users_qs = User.objects.filter(id=user.id, is_active=True).exclude(role="client")

        users = [
            {
                "id": row["id"],
                "name": row["first_name"] or row["username"],
                "monthly_threshold_hours": float(row["monthly_threshold_hours"] or 0),
            }
            for row in users_qs.values("id", "first_name", "username", "monthly_threshold_hours").order_by("first_name", "username")
        ]

        if user.is_superuser or user.role in ("superuser", "manager"):
            projects_qs = Project.objects.all()
        else:
            projects_qs = Project.objects.filter(projectassignment__assignee=user)

        projects = [
            {
                "id": row["id"],
                "name": row["name"],
                "monthly_threshold_hours": float(row["monthly_threshold_hours"] or 0),
            }
            for row in projects_qs.values("id", "name", "monthly_threshold_hours").order_by("name").distinct()
        ]

        from collections import defaultdict

        report = []
        per_user_seconds = defaultdict(float)
        per_project_seconds = defaultdict(float)

        grouped = qs.values("user_id", "user__first_name", "user__username", "project_id", "project__name", "billable").annotate(total=Sum("duration")).order_by("-total")
        merged_report: dict[tuple[int, int], dict[str, Any]] = {}

        for row in grouped:
            total = row["total"]
            if isinstance(total, int):
                total = timedelta(microseconds=total)
            seconds = float((total or timedelta(0)).total_seconds())

            key = (row["user_id"], row["project_id"])
            if key not in merged_report:
                merged_report[key] = {
                    "user_id": row["user_id"],
                    "user_name": row["user__first_name"] or row["user__username"],
                    "project_id": row["project_id"],
                    "project_name": row["project__name"],
                    "total_seconds": 0.0,
                    "billable_seconds": 0.0,
                    "non_billable_seconds": 0.0,
                }

            merged_report[key]["total_seconds"] += seconds
            if row["billable"]:
                merged_report[key]["billable_seconds"] += seconds
            else:
                merged_report[key]["non_billable_seconds"] += seconds
            per_user_seconds[row["user_id"]] += seconds
            per_project_seconds[row["project_id"]] += seconds

        for item in merged_report.values():
            report.append(
                {
                    "user_id": item["user_id"],
                    "user_name": item["user_name"],
                    "project_id": item["project_id"],
                    "project_name": item["project_name"],
                    "total_hours": round(item["total_seconds"] / 3600.0, 2),
                    "billable_hours": round(item["billable_seconds"] / 3600.0, 2),
                    "non_billable_hours": round(item["non_billable_seconds"] / 3600.0, 2),
                }
            )

        total_hours_overall = round(sum(per_user_seconds.values()) / 3600.0, 2)

        daily_rows = qs.annotate(day=TruncDate("start_datetime", tzinfo=tz)).values("day").annotate(total=Sum("duration")).order_by("day")
        daily_labels = [row["day"] for row in daily_rows]
        daily_hours = []
        for row in daily_rows:
            total = row["total"]
            if isinstance(total, int):
                total = timedelta(microseconds=total)
            daily_hours.append(round(((total or timedelta(0)).total_seconds()) / 3600.0, 2))

        dist_map = {}
        if qp_project_id and not qp_user_id:
            dist_rows = qs.values("user__first_name", "user__username").annotate(total=Sum("duration")).order_by("-total")
            for row in dist_rows:
                total = row["total"]
                if isinstance(total, int):
                    total = timedelta(microseconds=total)
                dist_map[row["user__first_name"] or row["user__username"]] = round((total or timedelta(0)).total_seconds() / 3600.0, 2)
        else:
            dist_rows = qs.values("project__name").annotate(total=Sum("duration")).order_by("-total")
            for row in dist_rows:
                total = row["total"]
                if isinstance(total, int):
                    total = timedelta(microseconds=total)
                dist_map[row["project__name"]] = round((total or timedelta(0)).total_seconds() / 3600.0, 2)

        tag_qs = qs.prefetch_related("tags")
        tag_totals = defaultdict(float)
        untagged = TagModel.objects.filter(name="Untagged", category="system").first()
        for entry in tag_qs:
            seconds = float((entry.duration or timedelta(0)).total_seconds())
            entry_tags = list(entry.tags.all())
            if entry_tags:
                share = seconds / len(entry_tags)
                for tag in entry_tags:
                    tag_totals[tag.id] += share
            elif untagged:
                tag_totals[untagged.id] += seconds

        tags_map = {tag.id: tag for tag in TagModel.objects.filter(id__in=tag_totals.keys())}
        tag_summary = []
        for tag_id, seconds in sorted(tag_totals.items(), key=lambda pair: pair[1], reverse=True):
            tag = tags_map.get(tag_id)
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            if tag:
                tag_summary.append(
                    {
                        "id": tag.id,
                        "name": tag.name,
                        "emoji": tag.emoji,
                        "category": tag.category,
                        "total_hours": hours,
                        "total_minutes": minutes,
                        "total_seconds": int(seconds),
                        "formatted": f"{hours}h {minutes:02d}m",
                    }
                )

        matrix_data = qs.values(
            "project_id",
            "project__name",
            "task_id",
            "task__title",
            "user_id",
            "user__first_name",
            "user__username",
        ).annotate(total=Sum("duration")).order_by("project__name", "task__title", "user__first_name")

        projects_map: dict[Any, dict[str, Any]] = {}
        for row in matrix_data:
            pid = row["project_id"]
            tid = row["task_id"]
            uid = row["user_id"]

            total = row["total"]
            if isinstance(total, int):
                total = timedelta(microseconds=total)
            seconds = float((total or timedelta(0)).total_seconds())
            hours = seconds / 3600.0

            if pid not in projects_map:
                projects_map[pid] = {"id": pid, "name": row["project__name"], "total_seconds": 0.0, "tasks": {}}
            project_node = projects_map[pid]
            project_node["total_seconds"] += seconds

            task_key = tid or -1
            task_title = row["task__title"] or "(No Task)"
            if task_key not in project_node["tasks"]:
                project_node["tasks"][task_key] = {"id": tid, "title": task_title, "total_seconds": 0.0, "users": []}
            task_node = project_node["tasks"][task_key]
            task_node["total_seconds"] += seconds

            user_name = row["user__first_name"] or row["user__username"]
            task_node["users"].append({"id": uid, "name": user_name, "total_hours": round(hours, 2)})

        task_matrix = []
        for _, project_data in sorted(projects_map.items(), key=lambda pair: pair[1]["name"]):
            tasks_list = []
            for _, task_data in sorted(project_data["tasks"].items(), key=lambda pair: pair[1]["title"]):
                tasks_list.append(
                    {
                        "id": task_data["id"],
                        "title": task_data["title"],
                        "total_hours": round(task_data["total_seconds"] / 3600.0, 2),
                        "users": task_data["users"],
                    }
                )
            task_matrix.append(
                {
                    "id": project_data["id"],
                    "name": project_data["name"],
                    "total_hours": round(project_data["total_seconds"] / 3600.0, 2),
                    "tasks": tasks_list,
                }
            )

        total_billable_seconds = sum(item["billable_seconds"] for item in merged_report.values())
        total_non_billable_seconds = sum(item["non_billable_seconds"] for item in merged_report.values())

        return Response(
            {
                "users": users,
                "projects": projects,
                "report": report,
                "daily": {"labels": daily_labels, "hours": daily_hours},
                "distribution": {"labels": list(dist_map.keys()), "hours": list(dist_map.values())},
                "totals": {
                    "overall_hours": total_hours_overall,
                    "billable_hours": round(total_billable_seconds / 3600.0, 2),
                    "non_billable_hours": round(total_non_billable_seconds / 3600.0, 2),
                    "per_user_hours": {str(k): round(v / 3600.0, 2) for k, v in per_user_seconds.items()},
                    "per_project_hours": {str(k): round(v / 3600.0, 2) for k, v in per_project_seconds.items()},
                },
                "tag_summary": tag_summary,
                "task_matrix": task_matrix,
            }
        )


def _resolve_date_range(request: Request, default_start: date, default_end: date) -> tuple[date, date] | Response:
    """Parse start_date/end_date (and a single `day` override) into a date
    window. Returns a Response on a bad format so callers can early-return."""
    start_param = _get_query_str(request, "start_date")
    end_param = _get_query_str(request, "end_date")
    day_param = _get_query_str(request, "day")
    if day_param:
        start_param = end_param = day_param

    start_date, end_date = default_start, default_end
    if start_param:
        try:
            start_date = datetime.fromisoformat(start_param).date()
        except ValueError:
            return Response({"detail": "Invalid start_date format"}, status=400)
    if end_param:
        try:
            end_date = datetime.fromisoformat(end_param).date()
        except ValueError:
            return Response({"detail": "Invalid end_date format"}, status=400)
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return start_date, end_date


class ProjectReportView(APIView):
    """Per-project rollup: tracked hours, task status breakdown, team size and
    budget vs. actual spend. Managers and superusers only."""

    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request: Request):
        tz = pytz.timezone("Asia/Kolkata")
        now_ist = timezone.now().astimezone(tz)
        default_start = now_ist.replace(day=1).date()
        default_end = (default_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        resolved = _resolve_date_range(request, default_start, default_end)
        if isinstance(resolved, Response):
            return resolved
        start_date, end_date = resolved

        start_dt = tz.localize(datetime.combine(start_date, time.min))
        end_dt = tz.localize(datetime.combine(end_date, time.max))

        projects = list(
            Project.objects.select_related("client").all().order_by("name")
        )
        project_ids = [p.id for p in projects]

        # Tracked hours within the window, per project.
        hours_map: dict[int, float] = {}
        billable_map: dict[int, float] = {}
        time_rows = (
            TimeEntry.objects.filter(
                project_id__in=project_ids,
                start_datetime__range=(start_dt, end_dt),
            )
            .values("project_id", "billable")
            .annotate(total=Sum("duration"))
        )
        for row in time_rows:
            total = row["total"]
            if isinstance(total, int):
                total = timedelta(microseconds=total)
            seconds = float((total or timedelta(0)).total_seconds())
            pid = row["project_id"]
            hours_map[pid] = hours_map.get(pid, 0.0) + seconds
            if row["billable"]:
                billable_map[pid] = billable_map.get(pid, 0.0) + seconds

        # Task counts grouped by status.
        task_map: dict[int, dict[str, int]] = {}
        for row in (
            Task.objects.filter(project_id__in=project_ids)
            .values("project_id", "status")
            .annotate(c=Count("id"))
        ):
            task_map.setdefault(row["project_id"], {})[row["status"]] = row["c"]

        # Distinct assigned team members per project.
        team_map: dict[int, int] = {
            row["project_id"]: row["c"]
            for row in ProjectAssignment.objects.filter(project_id__in=project_ids)
            .values("project_id")
            .annotate(c=Count("assignee", distinct=True))
        }

        # Budget envelope and actual recorded spend (expense ledger).
        budget_map: dict[int, float] = {
            b.project_id: float(b.planned_amount)
            for b in ProjectBudget.objects.filter(project_id__in=project_ids)
        }
        spend_map: dict[int, float] = {
            row["project_id"]: float(row["total"] or 0)
            for row in Transaction.objects.filter(
                flow="expense", project_id__in=project_ids,
            )
            .values("project_id")
            .annotate(total=Sum("amount"))
        }

        report = []
        for p in projects:
            tasks = task_map.get(p.id, {})
            total_tasks = sum(tasks.values())
            done_tasks = tasks.get("done", 0)
            total_seconds = hours_map.get(p.id, 0.0)
            billable_seconds = billable_map.get(p.id, 0.0)
            report.append(
                {
                    "id": p.id,
                    "name": p.name,
                    "client": p.client.name if p.client else None,
                    "billable": p.billable,
                    "start_date": p.start_date.isoformat() if p.start_date else None,
                    "end_date": p.end_date.isoformat() if p.end_date else None,
                    "due_date": p.due_date.isoformat() if p.due_date else None,
                    "total_hours": round(total_seconds / 3600.0, 2),
                    "billable_hours": round(billable_seconds / 3600.0, 2),
                    "monthly_threshold_hours": float(p.monthly_threshold_hours or 0),
                    "team_size": team_map.get(p.id, 0),
                    "tasks": {
                        "total": total_tasks,
                        "todo": tasks.get("todo", 0),
                        "pending": tasks.get("pending", 0),
                        "in_progress": tasks.get("in_progress", 0),
                        "done": done_tasks,
                        "completion_pct": round((done_tasks / total_tasks) * 100, 1) if total_tasks else 0.0,
                    },
                    "budget": {
                        "planned": budget_map.get(p.id, 0.0),
                        "spent": spend_map.get(p.id, 0.0),
                    },
                }
            )

        return Response(
            {
                "range": {"start": start_date.isoformat(), "end": end_date.isoformat()},
                "report": report,
                "totals": {
                    "projects": len(report),
                    "total_hours": round(sum(r["total_hours"] for r in report), 2),
                    "open_tasks": sum(r["tasks"]["total"] - r["tasks"]["done"] for r in report),
                },
            }
        )


class FinanceReportView(APIView):
    """Income vs. expense over time for the income/expense comparison graph.
    Executives (and superusers) only."""

    permission_classes = [IsAuthenticated, IsExecutive]

    def get(self, request: Request):
        today = timezone.localdate()
        # Default window: the trailing 12 months, anchored to month starts.
        default_end = today
        default_start = (today.replace(day=1) - timedelta(days=365)).replace(day=1)

        resolved = _resolve_date_range(request, default_start, default_end)
        if isinstance(resolved, Response):
            return resolved
        start_date, end_date = resolved

        in_range = Transaction.objects.filter(
            occurred_on__gte=start_date, occurred_on__lte=end_date,
        )

        # Monthly income/expense buckets, aligned on a shared month axis.
        monthly: dict[str, dict[str, float]] = {}
        for row in (
            in_range.annotate(m=TruncMonth("occurred_on"))
            .values("m", "flow")
            .annotate(total=Sum("amount"))
            .order_by("m")
        ):
            label = row["m"].strftime("%Y-%m")
            bucket = monthly.setdefault(label, {"income": 0.0, "expense": 0.0})
            bucket[row["flow"]] = float(row["total"] or 0)

        labels = sorted(monthly.keys())
        income_series = [round(monthly[m]["income"], 2) for m in labels]
        expense_series = [round(monthly[m]["expense"], 2) for m in labels]
        net_series = [round(i - e, 2) for i, e in zip(income_series, expense_series)]

        range_income = sum(income_series)
        range_expense = sum(expense_series)

        category_breakdown = [
            {"category": row["category"], "total": float(row["total"] or 0)}
            for row in in_range.filter(flow="expense")
            .values("category")
            .annotate(total=Sum("amount"))
            .order_by("-total")
        ]

        # All-time balance for the headline figures.
        all_income = Transaction.objects.filter(flow="income").aggregate(
            s=Sum("amount"))["s"] or 0
        all_expense = Transaction.objects.filter(flow="expense").aggregate(
            s=Sum("amount"))["s"] or 0

        return Response(
            {
                "currency": "INR",
                "range": {"start": start_date.isoformat(), "end": end_date.isoformat()},
                "series": {
                    "labels": labels,
                    "income": income_series,
                    "expense": expense_series,
                    "net": net_series,
                },
                "totals": {
                    "range_income": round(range_income, 2),
                    "range_expense": round(range_expense, 2),
                    "range_net": round(range_income - range_expense, 2),
                    "balance": float(all_income) - float(all_expense),
                    "income_total": float(all_income),
                    "expense_total": float(all_expense),
                },
                "category_breakdown": category_breakdown,
            }
        )


class ClockTimeReportView(APIView):
    """Worked-hours time report built from clock-in/out sessions.

    This is the "Time report of all users" — it reports actual attendance
    (ClockSession.worked_duration, i.e. gross minus lunch), not project
    timer entries. Supports a day-wise filter via `day`, or a
    `start_date`/`end_date` range. Visibility mirrors the rest of the app:
    employees see only themselves, managers their team, superusers everyone.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        user = cast(User, request.user)
        tz = pytz.timezone("Asia/Kolkata")
        now_ist = timezone.now().astimezone(tz)
        default_start = now_ist.replace(day=1).date()
        default_end = (default_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        resolved = _resolve_date_range(request, default_start, default_end)
        if isinstance(resolved, Response):
            return resolved
        start_date, end_date = resolved

        qs = ClockSession.objects.select_related("user").filter(
            date__gte=start_date, date__lte=end_date,
        )

        # Visibility scoping — identical rules to the project-time report.
        is_employee_only = (
            not user.is_superuser
            and getattr(user, "role", "") not in ("superuser", "manager")
        )
        if is_employee_only:
            qs = qs.filter(user_id=user.id)
        elif not (user.is_superuser or user.role == "superuser"):
            qs = qs.filter(user_id__in=build_visible_user_ids(user))

        qp_user_id = request.query_params.get("user_id")
        if qp_user_id and not is_employee_only:
            qs = qs.filter(user_id=user.id if qp_user_id == "me" else qp_user_id)

        now = timezone.now()
        per_user: dict[int, dict[str, Any]] = {}
        per_day: dict[str, float] = {}
        rows = []

        for sess in qs.order_by("clock_in_time"):
            # Open sessions (still clocked in) count up to "now" so today's
            # in-progress shift is reflected.
            end = sess.clock_out_time or now
            gross = (end - sess.clock_in_time).total_seconds() if sess.clock_in_time else 0.0
            lunch = sess.lunch_duration
            worked = gross - (lunch.total_seconds() if lunch else 0.0)
            worked = max(worked, 0.0)

            day_label = sess.date.isoformat()
            per_day[day_label] = per_day.get(day_label, 0.0) + worked

            uid = sess.user_id
            if uid not in per_user:
                per_user[uid] = {
                    "user_id": uid,
                    "name": sess.user.first_name or sess.user.username,
                    "seconds": 0.0,
                    "sessions": 0,
                }
            per_user[uid]["seconds"] += worked
            per_user[uid]["sessions"] += 1

            rows.append({
                "date": day_label,
                "user_id": uid,
                "user_name": sess.user.first_name or sess.user.username,
                "clock_in": sess.clock_in_time.astimezone(tz).strftime("%H:%M") if sess.clock_in_time else None,
                "clock_out": sess.clock_out_time.astimezone(tz).strftime("%H:%M") if sess.clock_out_time else None,
                "open": sess.clock_out_time is None,
                "worked_hours": round(worked / 3600.0, 2),
                "worked_seconds": int(round(worked)),
            })

        labels = sorted(per_day.keys())
        report = sorted(
            (
                {
                    "user_id": v["user_id"],
                    "user_name": v["name"],
                    "total_hours": round(v["seconds"] / 3600.0, 2),
                    "total_seconds": int(round(v["seconds"])),
                    "sessions": v["sessions"],
                }
                for v in per_user.values()
            ),
            key=lambda r: r["total_seconds"],
            reverse=True,
        )
        total_seconds = sum(v["seconds"] for v in per_user.values())

        return Response({
            "range": {"start": start_date.isoformat(), "end": end_date.isoformat()},
            "daily": {
                "labels": labels,
                "hours": [round(per_day[d] / 3600.0, 2) for d in labels],
                "seconds": [int(round(per_day[d])) for d in labels],
            },
            "report": report,
            "rows": rows,
            "totals": {
                "total_hours": round(total_seconds / 3600.0, 2),
                "total_seconds": int(round(total_seconds)),
                "users": len(per_user),
                "sessions": len(rows),
            },
        })
